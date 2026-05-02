from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.config import EXPORT_DIR
from services.class_service import ClassService
from services.grade_service import GradeService
from services.remedial_service import RemedialService
from services.report_service import ReportService
from services.student_service import StudentService
from services.subject_service import SubjectService


class ExcelService:
    def __init__(
        self,
        class_service: ClassService,
        subject_service: SubjectService,
        student_service: StudentService,
        grade_service: GradeService,
        remedial_service: RemedialService,
        report_service: ReportService,
        export_dir: Path | None = None,
        ) -> None:
        self.class_service = class_service
        self.subject_service = subject_service
        self.student_service = student_service
        self.grade_service = grade_service
        self.remedial_service = remedial_service
        self.report_service = report_service
        self.export_dir = export_dir or EXPORT_DIR

    def _target_dir(self, subfolder: str) -> Path:
        path = self.export_dir / subfolder
        path.mkdir(parents=True, exist_ok=True)
        return path

    def create_student_template(self) -> str:
        columns = [
            "nama_lengkap",
            "nis",
            "nisn",
            "kelas",
            "gender",
            "alamat",
            "nama_orang_tua",
            "nomor_wa",
        ]
        rows = [
            {
                "nama_lengkap": "Ahmad Fauzi",
                "nis": "1001",
                "nisn": "9988776655",
                "kelas": "VII A",
                "gender": "L",
                "alamat": "Jl. Melati No. 10",
                "nama_orang_tua": "Bapak Fauzan",
                "nomor_wa": "081234567890",
            },
            {
                "nama_lengkap": "Siti Aisyah",
                "nis": "1002",
                "nisn": "9988776656",
                "kelas": "VII A",
                "gender": "P",
                "alamat": "Jl. Kenanga No. 5",
                "nama_orang_tua": "Ibu Aminah",
                "nomor_wa": "081298765432",
            },
        ]
        frame = pd.DataFrame(rows, columns=columns)
        return self.export_dataframe(frame, "template_import_data_siswa", subfolder="template")

    def create_grade_template(self, subject_id: int | None = None) -> str:
        if subject_id:
            active_layout = self.grade_service.get_component_layout(subject_id)
            subject = self.subject_service.get_subject_by_id(subject_id)
            subject_name = subject["subject_name"] if subject else "Matematika"
        else:
            active_layout = [
                component
                for component in self.grade_service.get_component_blueprint(self.grade_service._get_default_component_scheme())
                if int(component["is_active"]) == 1
            ]
            subject_name = "Matematika"
        component_codes = [component["component_code"] for component in active_layout]
        columns = ["nis", "nama_lengkap", "kelas", "mata_pelajaran", *component_codes, "tambahan"]
        rows = [
            {
                "nis": "1001",
                "nama_lengkap": "Ahmad Fauzi",
                "kelas": "VII A",
                "mata_pelajaran": subject_name,
                **{code: value for code, value in zip(component_codes, [80, 78, 82, "", "", 76, 84, "", "", "", "", ""])},
                "tambahan": 0,
            },
            {
                "nis": "1002",
                "nama_lengkap": "Siti Aisyah",
                "kelas": "VII A",
                "mata_pelajaran": subject_name,
                **{code: value for code, value in zip(component_codes, [88, 90, 87, "", "", 86, 89, "", "", "", "", ""])},
                "tambahan": 0,
            },
        ]
        frame = pd.DataFrame(rows, columns=columns)
        return self.export_dataframe(frame, "template_import_nilai", subfolder="template")

    def validate_excel_columns(self, columns: list[str], required: list[str]) -> None:
        normalized = {column.strip().lower() for column in columns}
        if not set(required).issubset(normalized):
            raise ValueError("Format Excel tidak sesuai. Silakan gunakan template yang disediakan.")

    def import_students_excel(self, file_path: str) -> int:
        frame = pd.read_excel(file_path)
        frame.columns = [str(col).strip().lower() for col in frame.columns]
        required = ["nama_lengkap", "kelas"]
        self.validate_excel_columns(frame.columns.tolist(), required)
        classes = {row["class_name"]: row["id"] for row in self.class_service.get_classes()}
        imported = 0
        for _, row in frame.fillna("").iterrows():
            class_name = str(row.get("kelas", "")).strip()
            if not class_name:
                raise ValueError("Data wajib belum lengkap.")
            if class_name not in classes:
                self.class_service.add_class(class_name, "")
                classes = {item["class_name"]: item["id"] for item in self.class_service.get_classes()}
            self.student_service.add_student(
                {
                    "full_name": str(row.get("nama_lengkap", "")).strip(),
                    "nis": str(row.get("nis", "")).strip(),
                    "nisn": str(row.get("nisn", "")).strip(),
                    "gender": str(row.get("gender", "")).strip(),
                    "class_id": classes[class_name],
                    "address": str(row.get("alamat", "")).strip(),
                    "parent_name": str(row.get("nama_orang_tua", "")).strip(),
                    "parent_phone": str(row.get("nomor_wa", "")).strip(),
                }
            )
            imported += 1
        return imported

    def import_grades_excel(self, file_path: str) -> int:
        frame = pd.read_excel(file_path)
        frame.columns = [str(col).strip().lower() for col in frame.columns]
        required = ["kelas", "mata_pelajaran"]
        self.validate_excel_columns(frame.columns.tolist(), required)
        students = self.student_service.search_students()
        subjects = {row["subject_name"]: row["id"] for row in self.subject_service.get_subjects()}
        imported = 0
        for _, row in frame.fillna("").iterrows():
            subject_name = str(row.get("mata_pelajaran", "")).strip()
            class_name = str(row.get("kelas", "")).strip()
            if subject_name not in subjects:
                self.subject_service.add_subject(subject_name, "", None)
                subjects = {item["subject_name"]: item["id"] for item in self.subject_service.get_subjects()}
            subject_id = subjects[subject_name]
            active_layout = self.grade_service.get_component_layout(subject_id)
            active_component_codes = [
                component["component_code"]
                for component in active_layout
                if component["component_code"] not in {"uts", "uas"}
            ]
            match = next(
                (
                    item for item in students
                    if (
                        (
                            str(row.get("nis", "")).strip()
                            and item.get("nis") == str(row.get("nis", "")).strip()
                        )
                        or item["full_name"].lower() == str(row.get("nama_lengkap", "")).strip().lower()
                    )
                    and (not class_name or item.get("class_name") == class_name)
                ),
                None,
            )
            if not match:
                raise ValueError("Data siswa pada file nilai tidak ditemukan di database.")
            self.grade_service.save_grade(
                {
                    "student_id": match["id"],
                    "subject_id": subject_id,
                    "component_scores": {
                        **{
                            code: row.get(code, row.get("tugas", "") if active_component_codes and code == active_component_codes[0] else "")
                            for code in active_component_codes
                        },
                        **({"uts": row.get("uts", 0) or 0} if any(item["component_code"] == "uts" for item in active_layout) else {}),
                        **({"uas": row.get("uas", 0) or 0} if any(item["component_code"] == "uas" for item in active_layout) else {}),
                    },
                    "extra_score": row.get("tambahan", 0) or 0,
                }
            )
            imported += 1
        return imported

    def export_dataframe(
        self,
        frame: pd.DataFrame,
        filename: str,
        target_path: str | None = None,
        as_csv: bool = False,
        subfolder: str = "",
    ) -> str:
        base_dir = self._target_dir(subfolder) if subfolder else self.export_dir
        base_dir.mkdir(parents=True, exist_ok=True)
        path = Path(target_path) if target_path else base_dir / filename
        if as_csv:
            path = path.with_suffix(".csv")
            frame.to_csv(path, index=False)
        else:
            path = path.with_suffix(".xlsx")
            frame.to_excel(path, index=False)
        return str(path)

    def export_students_excel(self, class_id: int | None = None, as_csv: bool = False) -> str:
        data = self.student_service.search_students(class_id=class_id)
        if not data:
            raise ValueError("Data kosong.")
        frame = pd.DataFrame(data)
        frame = frame.rename(
            columns={
                "full_name": "Nama Siswa",
                "nis": "NIS",
                "nisn": "NISN",
                "class_name": "Kelas",
                "gender": "Gender",
                "address": "Alamat",
                "parent_name": "Nama Orang Tua",
                "parent_phone": "Nomor WA Orang Tua",
            }
        )
        frame.insert(0, "No", range(1, len(frame) + 1))
        return self.export_dataframe(frame, f"data_siswa_{datetime.now():%Y%m%d}", as_csv=as_csv, subfolder="data_siswa")

    def export_grades_excel(self, class_id: int | None = None, subject_id: int | None = None, as_csv: bool = False) -> str:
        if subject_id:
            rows = self.grade_service.get_grade_rows(class_id or 0, subject_id)
            if not rows:
                raise ValueError("Data kosong.")
            subject = self.subject_service.get_subject_by_id(subject_id)
            component_layout = self.grade_service.get_component_layout(subject_id)
            export_rows = []
            for row in rows:
                item = {
                    "Nama Siswa": row["full_name"],
                    "Kelas": row["class_name"],
                    "Mata Pelajaran": subject["subject_name"] if subject else "",
                }
                for component in component_layout:
                    item[component["component_name"]] = row["component_scores"].get(component["component_code"], 0.0)
                item["Rata-rata Harian"] = row["daily_score"]
                item["Tambahan"] = row["extra_score"]
                item["Nilai Akhir"] = row["final_result"]
                item["Predikat"] = row["predicate"]
                item["Status"] = row["status"]
                item["Ranking"] = row["rank_number"]
                export_rows.append(item)
            frame = pd.DataFrame(export_rows)
        else:
            data = self.grade_service.get_all_grades(class_id, subject_id)
            if not data:
                raise ValueError("Data kosong.")
            frame = pd.DataFrame(data)
            frame = frame.rename(
                columns={
                    "full_name": "Nama Siswa",
                    "nis": "NIS",
                    "class_name": "Kelas",
                    "subject_name": "Mata Pelajaran",
                    "task_score": "Rata-rata Harian",
                    "mid_score": "UTS",
                    "final_score": "UAS",
                    "extra_score": "Tambahan",
                    "final_result": "Nilai Akhir",
                    "predicate": "Predikat",
                    "status": "Status",
                    "rank_number": "Ranking",
                }
            )
        frame.insert(0, "No", range(1, len(frame) + 1))
        return self.export_dataframe(frame, f"nilai_{datetime.now():%Y%m%d}", as_csv=as_csv, subfolder="nilai")

    def export_reports_excel(self, class_id: int | None = None, subject_id: int | None = None, as_csv: bool = False) -> str:
        if not class_id:
            raise ValueError("Pilih kelas untuk membuat raport.")
        data = self.report_service.get_report_book_data(class_id)
        if not data:
            raise ValueError("Data kosong.")
        return self._create_report_workbook(
            data,
            f"raport_kelas_{datetime.now():%Y%m%d}",
            subfolder="raport",
        )

    def export_remedial_excel(self, class_id: int | None = None, subject_id: int | None = None, as_csv: bool = False) -> str:
        data = self.remedial_service.get_records(class_id, subject_id)
        if not data:
            raise ValueError("Data kosong.")
        frame = pd.DataFrame(data)
        frame = frame.rename(
            columns={
                "full_name": "Nama Siswa",
                "class_name": "Kelas",
                "subject_name": "Mapel",
                "original_score": "Nilai Awal",
                "recommended_score": "Rekomendasi",
                "remedial_score": "Nilai Remedial",
                "adjusted_score": "Nilai Setelah Remedial",
                "remedial_status": "Status Remedial",
                "notes": "Catatan",
            }
        )
        frame.insert(0, "No", range(1, len(frame) + 1))
        return self.export_dataframe(
            frame,
            f"smart_ketuntasan_{datetime.now():%Y%m%d}",
            as_csv=as_csv,
            subfolder="smart_ketuntasan",
        )

    def _create_report_workbook(self, students: list[dict], filename: str, *, subfolder: str) -> str:
        workbook = Workbook()
        workbook.remove(workbook.active)
        thin = Side(style="thin", color="666666")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill = PatternFill("solid", fgColor="D9EAF7")
        score_fill = PatternFill("solid", fgColor="E8F1FF")
        desc_fill = PatternFill("solid", fgColor="F8FBFF")
        for student in students:
            sheet_name = str(student["full_name"])[:31] or "Raport"
            ws = workbook.create_sheet(title=sheet_name)
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A11"
            widths = {"A": 5, "B": 28, "C": 14, "D": 12, "E": 58}
            for column, width in widths.items():
                ws.column_dimensions[column].width = width

            ws.merge_cells("A1:E1")
            ws["A1"] = ""
            ws.merge_cells("A2:E2")
            ws["A2"] = ""
            ws.merge_cells("A3:E3")
            ws["A3"] = "RAPORT HASIL BELAJAR"
            ws["A3"].font = Font(size=14, bold=True)
            ws["A3"].alignment = Alignment(horizontal="center")

            info_rows = [
                ("Nama", student["full_name"], "Kelas", student["class_name"]),
                ("NIS", student["nis"], "Semester", student["semester"]),
                ("NISN", student["nisn"], "Tahun Pelajaran", student["academic_year"]),
                ("Nama Orang Tua", student["parent_name"], "Wali Kelas", student["teacher_name"]),
            ]
            start_row = 5
            for index, (left_label, left_value, right_label, right_value) in enumerate(info_rows, start=start_row):
                ws[f"A{index}"] = left_label
                ws[f"B{index}"] = left_value
                ws[f"C{index}"] = right_label
                ws[f"D{index}"] = right_value
                ws[f"A{index}"].font = Font(bold=True)
                ws[f"C{index}"].font = Font(bold=True)
                ws.merge_cells(f"D{index}:E{index}")

            header_row = 10
            headers = ["No", "Mata Pelajaran", "Nilai Akhir", "Predikat", "Deskripsi Capaian"]
            for column, label in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=column, value=label)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = title_fill
                cell.border = border

            data_row = header_row + 1
            for number, lesson in enumerate(student["lessons"], start=1):
                values = [
                    number,
                    lesson["subject_name"],
                    lesson["final_result"],
                    lesson["predicate"],
                    lesson["description"],
                ]
                for column, value in enumerate(values, start=1):
                    cell = ws.cell(row=data_row, column=column, value=value)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="center" if column in {1, 3, 4} else "left",
                        vertical="top",
                        wrap_text=True,
                    )
                    if column == 3:
                        cell.fill = score_fill
                    elif column == 5:
                        cell.fill = desc_fill
                data_row += 1

            notes_row = data_row + 1
            ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=5)
            ws.cell(row=notes_row, column=1, value="Catatan Orang Tua / Wali")
            ws.cell(row=notes_row, column=1).font = Font(bold=True)
            for offset in range(1, 4):
                row_no = notes_row + offset
                ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=5)
                for col in range(1, 6):
                    ws.cell(row=row_no, column=col).border = border

            sign_row = notes_row + 6
            ws["A" + str(sign_row)] = "Mengetahui,"
            ws["A" + str(sign_row + 1)] = "Kepala Sekolah,"
            ws["C" + str(sign_row)] = ""
            ws["D" + str(sign_row)] = "Wali Kelas,"
            ws["A" + str(sign_row + 4)] = ""
            ws["D" + str(sign_row + 4)] = ""

            for row_index in range(11, max(data_row, notes_row + 3) + 1):
                ws.row_dimensions[row_index].height = 28

        target_dir = self._target_dir(subfolder)
        path = target_dir / f"{filename}.xlsx"
        workbook.save(path)
        return str(path)
