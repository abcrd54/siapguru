from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from app.database import DatabaseService
from app.license import FirebaseLicenseProvider, LocalLicenseProvider, get_license_provider
from app.update_checker import FirebaseUpdateChecker
from app.workspace import WorkspaceManager
from app.workspace_runtime import copy_workspace_seed, restore_workspace_from_backup, seed_workspace_settings
from services.backup_service import BackupService
from services.class_service import ClassService
from services.excel_service import ExcelService
from services.grade_service import GradeService
from services.remedial_service import RemedialService
from services.report_service import ReportService
from services.settings_service import SettingsService
from services.student_service import StudentService
from services.subject_service import SubjectService


class SiapGuruFlowTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TemporaryDirectory()
        root = Path(self.temp_dir.name)
        self.db = DatabaseService(root / "siapguru_test.db")
        self.db.init_database()
        self.settings = SettingsService(self.db)
        self.classes = ClassService(self.db)
        self.subjects = SubjectService(self.db)
        self.students = StudentService(self.db, self.classes)
        self.grades = GradeService(self.db, self.settings)
        self.remedial = RemedialService(self.db, self.settings, self.grades)
        self.reports = ReportService(self.db, self.grades, self.remedial)
        self.excel = ExcelService(
            self.classes,
            self.subjects,
            self.students,
            self.grades,
            self.remedial,
            self.reports,
            export_dir=root / "exports",
        )
        self.backup = BackupService(self.db, backup_dir=root / "backups")

    def tearDown(self) -> None:
        self.db.close_connection()
        self.temp_dir.cleanup()

    def test_end_to_end_flow(self) -> None:
        self.settings.update_settings(
            {
                "school_name": "Sekolah Maju",
                "teacher_name": "Ibu Rina",
                "academic_year": "2025/2026",
                "semester": "Ganjil",
                "kkm": 75,
                "weight_task": 30,
                "weight_mid": 30,
                "weight_final": 40,
            }
        )

        self.classes.add_class("VII A", "Pak Dedi")
        class_row = self.classes.get_classes()[0]
        self.subjects.add_subject("Matematika", "Bu Sari", 78)
        subject_row = self.subjects.get_subjects()[0]

        self.students.add_student({"full_name": "Ani", "nis": "001", "class_id": class_row["id"]})
        self.students.add_student({"full_name": "Budi", "nis": "002", "class_id": class_row["id"]})
        student_rows = self.students.get_students_by_class(class_row["id"])
        ani = next(row for row in student_rows if row["full_name"] == "Ani")
        budi = next(row for row in student_rows if row["full_name"] == "Budi")

        self.grades.save_grade(
            {
                "student_id": ani["id"],
                "subject_id": subject_row["id"],
                "daily_scores": [70, 72, 68, None, None],
                "mid_score": 70,
                "final_score": 70,
                "extra_score": 0,
            }
        )
        self.grades.save_grade(
            {
                "student_id": budi["id"],
                "subject_id": subject_row["id"],
                "daily_scores": [90, 92, 88, None, None],
                "mid_score": 90,
                "final_score": 90,
                "extra_score": 0,
            }
        )
        self.grades.calculate_ranking(class_row["id"], subject_row["id"])

        all_grades = self.grades.get_all_grades(class_row["id"], subject_row["id"])
        self.assertEqual(len(all_grades), 2)
        all_grades_by_name = {row["full_name"]: row for row in all_grades}
        self.assertEqual(all_grades_by_name["Budi"]["rank_number"], 1)
        self.assertEqual(all_grades_by_name["Ani"]["rank_number"], 2)

        below = self.remedial.get_below_kkm_students(class_row["id"], subject_row["id"])
        self.assertEqual(len(below), 1)
        self.assertEqual(below[0]["full_name"], "Ani")
        self.remedial.apply_recommendation(below[0]["grade_id"], 78)

        grade_after = self.grades.get_all_grades(class_row["id"], subject_row["id"])
        ani_after = next(row for row in grade_after if row["full_name"] == "Ani")
        self.assertEqual(ani_after["status"], "Tuntas")

        generated = self.reports.generate_all_descriptions(class_row["id"], subject_row["id"])
        ani_report = next(row for row in generated if row["full_name"] == "Ani")
        self.assertEqual(ani_report["predicate"], "C")
        self.assertIn("Matematika", ani_report["description"])
        self.reports.save_description(
            ani_report["student_id"],
            subject_row["id"],
            ani_report["grade_id"],
            ani_report["description"],
        )
        saved = self.reports.get_saved_descriptions(class_row["id"], subject_row["id"])
        self.assertEqual(len(saved), 1)

        export_path = self.excel.export_grades_excel(class_row["id"], subject_row["id"])
        self.assertTrue(Path(export_path).exists())
        backup_path = self.backup.backup_database()
        self.assertTrue(Path(backup_path).exists())

    def test_validation_errors(self) -> None:
        with self.assertRaises(ValueError):
            self.settings.update_settings(
                {
                    "school_name": "",
                    "teacher_name": "",
                    "academic_year": "",
                    "semester": "Ganjil",
                    "kkm": 75,
                    "weight_task": 20,
                    "weight_mid": 20,
                    "weight_final": 20,
                }
            )

        self.classes.add_class("VII A", "")
        class_row = self.classes.get_classes()[0]
        with self.assertRaises(ValueError):
            self.students.add_student({"full_name": "", "class_id": class_row["id"]})

    def test_student_import_creates_missing_class(self) -> None:
        import pandas as pd

        import_file = Path(self.temp_dir.name) / "students.xlsx"
        pd.DataFrame(
            [
                {
                    "nama_lengkap": "Citra",
                    "nis": "003",
                    "nisn": "123",
                    "kelas": "VIII A",
                    "gender": "P",
                    "alamat": "",
                    "nama_orang_tua": "",
                    "nomor_wa": "",
                }
            ]
        ).to_excel(import_file, index=False)

        count = self.excel.import_students_excel(str(import_file))
        self.assertEqual(count, 1)
        classes = self.classes.get_classes()
        self.assertEqual(classes[0]["class_name"], "VIII A")

    def test_app_mode_can_be_stored(self) -> None:
        self.settings.set_app_mode("guru")
        self.assertEqual(self.settings.get_app_mode(), "guru")
        with self.assertRaises(ValueError):
            self.settings.set_app_mode("wali")

    def test_active_context_can_be_stored(self) -> None:
        self.classes.add_class("VII A", "")
        self.subjects.add_subject("Matematika", "", 77)
        class_row = self.classes.get_classes()[0]
        subject_row = self.subjects.get_subjects()[0]
        self.settings.update_active_context(
            default_class_id=class_row["id"],
            default_subject_id=subject_row["id"],
            primary_class_id=class_row["id"],
        )
        context = self.settings.get_active_context()
        self.assertEqual(context["default_class_id"], class_row["id"])
        self.assertEqual(context["default_subject_id"], subject_row["id"])
        self.assertEqual(context["primary_class_id"], class_row["id"])

    def test_workspace_manager_separates_mode_paths(self) -> None:
        root = Path(self.temp_dir.name) / "workspace-root"
        state = Path(self.temp_dir.name) / "state.json"
        manager = WorkspaceManager(root, state)
        manager.set_enabled_modes(["guru"])
        self.assertEqual(manager.get_enabled_modes(), ["guru"])
        manager.set_active_mode("guru")
        self.assertEqual(manager.get_active_mode(), "guru")
        self.assertTrue(manager.is_mode_enabled("guru"))

    def test_workspace_manager_creates_period_workspaces(self) -> None:
        root = Path(self.temp_dir.name) / "workspace-root"
        state = Path(self.temp_dir.name) / "state.json"
        manager = WorkspaceManager(root, state)
        manager.set_enabled_modes(["guru"])
        manager.save_profile(school_name="Sekolah Maju", teacher_name="Ibu Rina")
        workspace = manager.create_workspace(academic_year="2026/2027", semester="Ganjil")
        self.assertEqual(workspace["label"], "2026/2027 - Ganjil")
        self.assertEqual(manager.get_active_workspace_id(), workspace["id"])
        self.assertEqual(manager.list_workspaces()[0]["academic_year"], "2026/2027")
        self.assertTrue(str(manager.get_db_path(workspace["id"])).endswith("siapguru.db"))

    def test_workspace_copy_seed_can_copy_selected_master_data(self) -> None:
        source_db = DatabaseService(Path(self.temp_dir.name) / "source.db")
        source_db.init_database()
        source_settings = SettingsService(source_db)
        seed_workspace_settings(
            source_db,
            {"academic_year": "2025/2026", "semester": "Ganjil"},
            {"school_name": "Sekolah Maju", "teacher_name": "Ibu Rina"},
        )
        source_classes = ClassService(source_db)
        source_subjects = SubjectService(source_db)
        source_students = StudentService(source_db, source_classes)
        source_grades = GradeService(source_db, source_settings)
        source_classes.add_class("VII A", "Pak Dedi")
        class_row = source_classes.get_classes()[0]
        source_students.add_student({"full_name": "Ani", "nis": "001", "class_id": class_row["id"]})
        source_subjects.add_subject("Matematika", "Bu Sari", 78, 20, 30, 50)
        subject_row = source_subjects.get_subjects()[0]
        source_grades.update_component_layout(
            subject_row["id"],
            {
                "use_daily_components": 1,
                "daily_component_count": 2,
                "use_mid_component": 1,
                "use_final_component": 0,
            },
        )

        target_db = DatabaseService(Path(self.temp_dir.name) / "target.db")
        target_db.init_database()
        seed_workspace_settings(
            target_db,
            {"academic_year": "2026/2027", "semester": "Ganjil"},
            {"school_name": "Sekolah Maju", "teacher_name": "Ibu Rina"},
        )
        copy_workspace_seed(
            source_db,
            target_db,
            {
                "classes": True,
                "students": True,
                "subjects": True,
                "subject_rules": True,
                "assessment_components": True,
            },
        )
        copied_classes = ClassService(target_db).get_classes()
        copied_students = StudentService(target_db, ClassService(target_db)).get_students_by_class(copied_classes[0]["id"])
        copied_subjects = SubjectService(target_db).get_subjects()
        copied_layout = GradeService(target_db, SettingsService(target_db)).get_component_layout(copied_subjects[0]["id"])
        self.assertEqual(copied_classes[0]["class_name"], "VII A")
        self.assertEqual(copied_students[0]["full_name"], "Ani")
        self.assertEqual(copied_subjects[0]["weight_final"], 50)
        self.assertEqual([row["component_code"] for row in copied_layout], ["harian_1", "harian_2", "uts"])
        source_db.close_connection()
        target_db.close_connection()

    def test_local_license_provider_requires_activation(self) -> None:
        license_path = Path(self.temp_dir.name) / "license_local.json"
        keys_path = Path(self.temp_dir.name) / "license_keys_local.json"
        provider = LocalLicenseProvider(license_path, keys_path)
        profile = provider.get_profile()
        self.assertEqual(profile.source, "local")
        self.assertFalse(profile.is_activated)
        self.assertEqual(profile.enabled_modes, [])
        self.assertTrue(license_path.exists())
        self.assertTrue(keys_path.exists())

        activated = provider.activate_key("SG-GURU-001")
        self.assertTrue(activated.is_activated)
        self.assertIn("guru", activated.enabled_modes)
        self.assertEqual(activated.enabled_modes, ["guru"])

    def test_firebase_license_feature_mapping(self) -> None:
        provider = FirebaseLicenseProvider(cache_path=Path(self.temp_dir.name) / "firebase_cache.json")
        self.assertEqual(
            provider._extract_enabled_modes(
                {
                    "license_role": "guru",
                    "features": {"guru": True},
                }
            ),
            ["guru"],
        )

    def test_firebase_license_accepts_document_id_as_key(self) -> None:
        provider = FirebaseLicenseProvider(cache_path=Path(self.temp_dir.name) / "firebase_cache.json")
        provider._validate_license_document(
            {
                "docId": "SG-GURU-DOC-ID",
                "app_id": "siapguru",
                "status": "active",
            },
            "SG-GURU-DOC-ID",
        )

    def test_runtime_license_provider_is_firebase_only(self) -> None:
        provider = get_license_provider("firebase")
        self.assertIsInstance(provider, FirebaseLicenseProvider)
        with self.assertRaises(RuntimeError):
            get_license_provider("local")

    def test_update_checker_compares_versions_and_force_update(self) -> None:
        checker = FirebaseUpdateChecker(
            cache_path=Path(self.temp_dir.name) / "update_cache.json",
            credentials_path="",
            project_id="demo-project",
            collection_name="app_releases",
            channel="production",
        )
        optional = {
            "latest_version": "1.0.1",
            "minimum_version": "1.0.0",
            "force_update": False,
            "download_url": "https://siapdigital.web.id/download/siapguru-latest",
            "notes": ["Perbaikan minor"],
        }
        required = {
            "latest_version": "1.0.1",
            "minimum_version": "1.0.1",
            "force_update": True,
            "download_url": "https://siapdigital.web.id/download/siapguru-latest",
            "notes": ["Update wajib"],
        }
        checker._fetch_release_document = lambda: optional
        optional_result = checker.check("1.0.0")
        self.assertIsNotNone(optional_result)
        self.assertTrue(optional_result.update_available)
        self.assertFalse(optional_result.update_required)
        checker._fetch_release_document = lambda: required
        required_result = checker.check("1.0.0")
        self.assertIsNotNone(required_result)
        self.assertTrue(required_result.update_available)
        self.assertTrue(required_result.update_required)

    def test_update_checker_uses_cached_release_when_fetch_fails(self) -> None:
        cache_path = Path(self.temp_dir.name) / "update_cache.json"
        checker = FirebaseUpdateChecker(
            cache_path=cache_path,
            credentials_path="",
            project_id="demo-project",
            collection_name="app_releases",
            channel="production",
        )
        cache_path.write_text(
            """
            {
              "latest_version": "1.0.2",
              "minimum_version": "1.0.1",
              "force_update": true,
              "download_url": "https://siapdigital.web.id/download/siapguru-latest",
              "notes": ["Update wajib dari cache"]
            }
            """.strip(),
            encoding="utf-8",
        )
        checker._fetch_release_document = lambda: (_ for _ in ()).throw(RuntimeError("offline"))
        result = checker.check("1.0.0")
        self.assertIsNotNone(result)
        self.assertTrue(result.update_required)
        self.assertEqual(result.latest_version, "1.0.2")

    def test_subject_kkm_and_daily_average_are_used(self) -> None:
        self.classes.add_class("VII A", "")
        self.subjects.add_subject("IPA", "Pak Anton", 80)
        class_row = self.classes.get_classes()[0]
        subject_row = self.subjects.get_subjects()[0]
        self.students.add_student({"full_name": "Kiki", "nis": "009", "class_id": class_row["id"]})
        student = self.students.get_students_by_class(class_row["id"])[0]
        self.grades.save_grade(
            {
                "student_id": student["id"],
                "subject_id": subject_row["id"],
                "daily_scores": [70, 80, 90, None, None],
                "mid_score": 80,
                "final_score": 80,
                "extra_score": 0,
            }
        )
        rows = self.grades.get_grade_rows(class_row["id"], subject_row["id"])
        self.assertEqual(rows[0]["daily_score"], 80.0)
        self.assertEqual(rows[0]["kkm"], 80)
        self.assertEqual(rows[0]["status"], "Tuntas")

    def test_subject_specific_weights_override_default_settings(self) -> None:
        self.settings.update_settings(
            {
                "school_name": "Sekolah Maju",
                "teacher_name": "Ibu Rina",
                "academic_year": "2025/2026",
                "semester": "Ganjil",
                "weight_task": 30,
                "weight_mid": 30,
                "weight_final": 40,
            }
        )
        self.classes.add_class("VII A", "")
        self.subjects.add_subject("Informatika", "Bu Lina", 75, 20, 30, 50)
        class_row = self.classes.get_classes()[0]
        subject_row = self.subjects.get_subjects()[0]
        self.students.add_student({"full_name": "Rafi", "nis": "010", "class_id": class_row["id"]})
        student = self.students.get_students_by_class(class_row["id"])[0]
        self.grades.save_grade(
            {
                "student_id": student["id"],
                "subject_id": subject_row["id"],
                "daily_scores": [100, 100, 100],
                "mid_score": 50,
                "final_score": 50,
                "extra_score": 0,
            }
        )
        grade_row = self.grades.get_all_grades(class_row["id"], subject_row["id"])[0]
        self.assertEqual(grade_row["final_result"], 60.0)

    def test_subject_component_layout_affects_final_score(self) -> None:
        self.subjects.add_subject("IPS", "Bu Rara", 76, 20, 30, 50)
        subject_id = self.subjects.get_subjects()[0]["id"]
        self.grades.update_component_layout(
            subject_id,
            {
                "use_daily_components": 1,
                "daily_component_count": 2,
                "use_mid_component": 1,
                "use_final_component": 0,
            },
        )
        result = self.grades.calculate_final_score(80, 70, 10, 0, subject_id=subject_id)
        self.assertEqual(result, 74.0)

    def test_assessment_scheme_changes_component_layout(self) -> None:
        self.settings.update_settings(
            {
                "app_mode": "",
                "default_class_id": None,
                "default_subject_id": None,
                "primary_class_id": None,
                "school_name": "",
                "teacher_name": "",
                "academic_year": "",
                "semester": "Ganjil",
                "kkm": 75,
                "weight_task": 30,
                "weight_mid": 30,
                "weight_final": 40,
                "daily_component_count": 2,
            }
        )
        self.subjects.add_subject("Bahasa Indonesia", "Bu Dina", 78)
        subject_id = self.subjects.get_subjects()[0]["id"]
        layout = self.grades.get_component_layout(subject_id)
        codes = [item["component_code"] for item in layout]
        self.assertEqual(codes, ["harian_1", "harian_2", "uts", "uas"])

    def test_export_grades_includes_active_components(self) -> None:
        import pandas as pd

        self.settings.update_settings(
            {
                "app_mode": "guru",
                "default_class_id": None,
                "default_subject_id": None,
                "primary_class_id": None,
                "school_name": "Sekolah Maju",
                "teacher_name": "Ibu Rina",
                "academic_year": "2026/2027",
                "semester": "Ganjil",
                "kkm": 75,
                "weight_task": 30,
                "weight_mid": 30,
                "weight_final": 40,
                "daily_component_count": 2,
            }
        )
        self.classes.add_class("VII A", "")
        self.subjects.add_subject("IPA", "Pak Anton", 80)
        class_row = self.classes.get_classes()[0]
        subject_row = self.subjects.get_subjects()[0]
        self.students.add_student({"full_name": "Kiki", "nis": "009", "class_id": class_row["id"]})
        student = self.students.get_students_by_class(class_row["id"])[0]
        self.grades.save_grade(
            {
                "student_id": student["id"],
                "subject_id": subject_row["id"],
                "component_scores": {
                    "harian_1": 70,
                    "harian_2": 80,
                    "uts": 80,
                    "uas": 85,
                },
                "extra_score": 0,
            }
        )
        export_path = self.excel.export_grades_excel(class_row["id"], subject_row["id"])
        frame = pd.read_excel(export_path)
        self.assertIn("H1", frame.columns)
        self.assertIn("H2", frame.columns)
        self.assertIn("UTS", frame.columns)
        self.assertIn("UAS", frame.columns)

    def test_output_files_are_grouped_by_action_folder(self) -> None:
        from openpyxl import load_workbook

        self.classes.add_class("VII A", "")
        self.subjects.add_subject("IPA", "Pak Anton", 80)
        class_row = self.classes.get_classes()[0]
        subject_row = self.subjects.get_subjects()[0]
        self.students.add_student({"full_name": "Kiki", "nis": "009", "class_id": class_row["id"]})
        student = self.students.get_students_by_class(class_row["id"])[0]
        self.grades.save_grade(
            {
                "student_id": student["id"],
                "subject_id": subject_row["id"],
                "daily_scores": [70, 80, 90],
                "mid_score": 80,
                "final_score": 85,
                "extra_score": 0,
            }
        )
        self.reports.save_description(
            student["id"],
            subject_row["id"],
            1,
            "Deskripsi uji coba",
        )

        student_template = Path(self.excel.create_student_template())
        grade_template = Path(self.excel.create_grade_template())
        students_export = Path(self.excel.export_students_excel(class_row["id"]))
        grades_export = Path(self.excel.export_grades_excel(class_row["id"], subject_row["id"]))
        reports_export = Path(self.excel.export_reports_excel(class_row["id"], subject_row["id"]))
        backup_export = Path(self.backup.backup_database())

        self.assertEqual(student_template.parent.name, "template")
        self.assertEqual(grade_template.parent.name, "template")
        self.assertEqual(students_export.parent.name, "data_siswa")
        self.assertEqual(grades_export.parent.name, "nilai")
        self.assertEqual(reports_export.parent.name, "raport")
        self.assertEqual(backup_export.parent.name, "backups")
        workbook = load_workbook(reports_export)
        self.assertIn("Kiki", workbook.sheetnames)
        self.assertEqual(workbook["Kiki"]["A3"].value, "RAPORT HASIL BELAJAR")

    def test_backup_metadata_can_restore_as_new_workspace(self) -> None:
        root = Path(self.temp_dir.name)
        self.settings.update_settings(
            {
                "app_mode": "guru",
                "school_name": "Sekolah Maju",
                "teacher_name": "Ibu Rina",
                "academic_year": "2025/2026",
                "semester": "Ganjil",
            }
        )
        backup_path = Path(self.backup.backup_database())
        metadata = BackupService.read_backup_metadata(str(backup_path))
        self.assertEqual(metadata["academic_year"], "2025/2026")
        self.assertEqual(metadata["semester"], "Ganjil")
        self.assertEqual(metadata["teacher_name"], "Ibu Rina")
        self.assertTrue(metadata["backup_date"])

        manager = WorkspaceManager(root / "workspace-root", root / "state.json")
        manager.set_enabled_modes(["guru"])
        manager.save_profile(school_name="Sekolah Maju", teacher_name="Ibu Rina")
        restored_workspace = restore_workspace_from_backup(manager, str(backup_path))
        restored_db = DatabaseService(manager.get_db_path(restored_workspace["id"]))
        restored_db.init_database()
        restored_settings = SettingsService(restored_db).get_settings()
        self.assertEqual(restored_settings["academic_year"], "2025/2026")
        self.assertEqual(restored_settings["semester"], "Ganjil")
        restored_db.close_connection()

    def test_report_book_uses_effective_latest_grade_data(self) -> None:
        self.settings.update_settings(
            {
                "school_name": "Sekolah Maju",
                "teacher_name": "Ibu Rina",
                "academic_year": "2025/2026",
                "semester": "Ganjil",
                "weight_task": 30,
                "weight_mid": 30,
                "weight_final": 40,
            }
        )
        self.classes.add_class("VII A", "")
        class_id = self.classes.get_classes()[0]["id"]
        self.subjects.add_subject("Matematika", "Bu Sari", 78)
        subject_id = self.subjects.get_subjects()[0]["id"]
        self.students.add_student({"full_name": "Ani", "nis": "001", "class_id": class_id})
        student_id = self.students.get_students_by_class(class_id)[0]["id"]
        self.grades.save_grade(
            {
                "student_id": student_id,
                "subject_id": subject_id,
                "daily_scores": [70, 70, 70],
                "mid_score": 70,
                "final_score": 70,
                "extra_score": 0,
            }
        )
        grade_id = self.grades.get_all_grades(class_id, subject_id)[0]["id"]
        self.reports.save_description(student_id, subject_id, grade_id, "Deskripsi lama yang harus diperbarui")
        self.remedial.apply_recommendation(
            grade_id,
            80,
            "Penyesuaian uji",
            category="Auto Ketuntasan",
            status_label="Tuntas Setelah Penyesuaian",
        )
        report_book = self.reports.get_report_book_data(class_id)
        lesson = report_book[0]["lessons"][0]
        self.assertEqual(lesson["final_result"], 80.0)
        self.assertEqual(lesson["predicate"], "B")
        self.assertEqual(lesson["status"], "Tuntas Setelah Penyesuaian")
        self.assertIn("penyesuaian", lesson["description"].lower())

    def test_smart_ketuntasan_scales_scores_to_target_range(self) -> None:
        self.classes.add_class("VII A", "")
        self.subjects.add_subject("Matematika", "Bu Sari", 78)
        class_id = self.classes.get_classes()[0]["id"]
        subject_id = self.subjects.get_subjects()[0]["id"]
        self.students.add_student({"full_name": "Ani", "nis": "001", "class_id": class_id})
        self.students.add_student({"full_name": "Budi", "nis": "002", "class_id": class_id})
        ani, budi = self.students.get_students_by_class(class_id)

        self.grades.save_grade(
            {
                "student_id": ani["id"],
                "subject_id": subject_id,
                "daily_scores": [74, 75, 76],
                "mid_score": 77,
                "final_score": 78,
                "extra_score": 0,
            }
        )
        self.grades.save_grade(
            {
                "student_id": budi["id"],
                "subject_id": subject_id,
                "daily_scores": [88, 90, 92],
                "mid_score": 90,
                "final_score": 90,
                "extra_score": 0,
            }
        )

        candidates = self.remedial.get_adjustment_candidates(
            class_id,
            subject_id,
            target_min_score=75,
            target_max_score=96,
        )
        self.assertEqual(len(candidates), 2)
        self.assertEqual(candidates[0]["full_name"], "Ani")
        self.assertEqual(candidates[0]["adjusted_score"], 75.0)
        self.assertEqual(candidates[1]["full_name"], "Budi")
        self.assertEqual(candidates[1]["adjusted_score"], 96.0)

        applied = self.remedial.apply_bulk_adjustments(
            class_id,
            subject_id,
            target_min_score=75,
            target_max_score=96,
        )
        self.assertEqual(applied, 2)
        final_rows = self.grades.get_all_grades(class_id, subject_id)
        by_name = {row["full_name"]: row for row in final_rows}
        self.assertEqual(by_name["Ani"]["final_result"], 75.0)
        self.assertEqual(by_name["Budi"]["final_result"], 96.0)

    def test_grade_template_follows_subject_component_layout(self) -> None:
        import pandas as pd

        self.subjects.add_subject("IPS", "Bu Rara", 76)
        subject_id = self.subjects.get_subjects()[0]["id"]
        self.grades.update_component_layout(
            subject_id,
            {
                "use_daily_components": 1,
                "daily_component_count": 2,
                "use_mid_component": 1,
                "use_final_component": 0,
            },
        )
        template_path = self.excel.create_grade_template(subject_id)
        frame = pd.read_excel(template_path)
        self.assertIn("H1", frame.columns)
        self.assertIn("H2", frame.columns)
        self.assertIn("UTS", frame.columns)
        self.assertNotIn("UAS", frame.columns)

    def test_smart_ketuntasan_modes_follow_spec(self) -> None:
        kkm = 75
        tuntas = self.remedial.generate_recommendation(80, kkm, mode="minimal")
        self.assertEqual(tuntas["final_score"], 80)
        self.assertEqual(tuntas["category"], "Tuntas")
        self.assertFalse(tuntas["auto_applied"])

        minimal = self.remedial.generate_recommendation(72, kkm, mode="minimal")
        self.assertEqual(minimal["gap"], 3)
        self.assertEqual(minimal["final_score"], 75)
        self.assertEqual(minimal["category"], "Remedial Ringan")
        self.assertEqual(minimal["status"], "Auto Tuntas")
        self.assertTrue(minimal["auto_applied"])

        natural_small_gap = self.remedial.generate_recommendation(74, kkm, mode="natural")
        self.assertEqual(natural_small_gap["final_score"], 76)

        natural_medium_gap = self.remedial.generate_recommendation(71, kkm, mode="natural")
        self.assertEqual(natural_medium_gap["final_score"], 75)

        auto_far_gap = self.remedial.generate_recommendation(60, kkm, mode="minimal")
        self.assertEqual(auto_far_gap["final_score"], 75)
        self.assertEqual(auto_far_gap["category"], "Auto Ketuntasan")
        self.assertEqual(auto_far_gap["status"], "Auto Tuntas")
        self.assertTrue(auto_far_gap["auto_applied"])


if __name__ == "__main__":
    unittest.main()
